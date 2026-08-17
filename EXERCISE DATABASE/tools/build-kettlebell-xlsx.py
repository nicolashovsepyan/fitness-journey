#!/usr/bin/env python3
"""
KETTLEBELL.xlsx — the review sheet for the researched bell library.

  node "EXERCISE DATABASE/tools/build-kettlebell.mjs"      # writes the merge plan
  python3 "EXERCISE DATABASE/tools/build-kettlebell-xlsx.py"

Standalone on purpose: it does NOT touch EXERCISES.xlsx, which Nicolas has open.
Once he has said yes to the list, the keepers merge into MOVEMENTS.
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT  = os.path.join(HERE, '..', 'workbook', 'KETTLEBELL.xlsx')

HEAD_BG, HEAD_FG = '16213E', 'FFFFFF'
NEEDS = PatternFill('solid', fgColor='FFF2CC')
AUTO  = PatternFill('solid', fgColor='E8F5E9')
WARN  = PatternFill('solid', fgColor='FBE0DE')
THIN  = Border(*[Side(style='thin', color='D8D8D8')]*4)

KB    = json.load(open(os.path.join(HERE, 'kettlebell-library.json')))
MERGE = json.load(open(os.path.join(HERE, '..', 'workbook', '_kettlebell_merge.json')))

F = ['id','name','patterns','muscles_primary','muscles_secondary','laterality','measure',
     'level','diff','bells','sec_per_rep','contra','cue','standard']
rows = [dict(zip(F, a)) for a in KB['movements']]
fund = set(KB['kb_fundamental'])
merge_auto = {a: b for a, b, *_ in MERGE['likely']}
merge_auto.update({x[0]: x[1] for x in MERGE['exact']})
merge_auto.update({v: k for k, v in KB['merge_into_existing'].items()} and
                  {k: v for k, v in KB['merge_into_existing'].items()})

track_of = {}
for t, members in KB['tracks'].items():
    for i, mid in enumerate(members, 1):
        track_of[mid] = (t, i)


def head(ws, headers, row=1):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = PatternFill('solid', fgColor=HEAD_BG)
        c.font = Font(bold=True, color=HEAD_FG, size=10)
        c.alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 28


def sheet_movements(wb):
    ws = wb.create_sheet('KETTLEBELL')
    ws.cell(row=1, column=1, value=(
        '62 kettlebell movements, researched from StrongFirst / Pavel, Dan John, Onnit / Eric Leija '
        'and girevoy sources. Say yes / no / later in column C. The 10 marked FUNDAMENTAL are the '
        'cross-school consensus — I would keep all ten.')).font = Font(italic=True, size=10)
    ws.merge_cells('A1:P1')
    cols = ['id','Movement','Include?','Fundamental','Already in DB','Track','Rank','Patterns',
            'Primary muscles','Laterality','Measure','Level','Diff','Bells','Sec/rep',
            'Hard on (joints)','Cue','What counts as a rep']
    head(ws, cols, row=3)

    for j, r in enumerate(rows, start=4):
        dbid = merge_auto.get(r['id'], '')
        t, rank = track_of.get(r['id'], ('', ''))
        vals = [r['id'], r['name'], '', 'YES' if r['id'] in fund else '',
                dbid or 'new', t, rank, r['patterns'], r['muscles_primary'], r['laterality'],
                r['measure'], r['level'], r['diff'], r['bells'], r['sec_per_rep'] or '',
                r['contra'], r['cue'], r['standard']]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=j, column=i, value=v)
            c.border = THIN
            c.alignment = Alignment(vertical='top', wrap_text=(i in (17, 18)))
            if i == 3: c.fill = NEEDS
            if i == 4 and v == 'YES': c.fill = AUTO; c.font = Font(bold=True)
            if i == 5: c.fill = AUTO if dbid else PatternFill()
    last = 3 + len(rows)

    d = DataValidation(type='list', formula1='"yes,no,later"', allow_blank=True)
    ws.add_data_validation(d); d.add(f'C4:C{last}')
    ws.freeze_panes = 'C4'
    ws.auto_filter.ref = f'A3:R{last}'
    for i, w in enumerate([24, 30, 10, 12, 20, 16, 6, 30, 26, 12, 9, 7, 6, 7, 8, 24, 54, 54], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def sheet_tracks(wb):
    ws = wb.create_sheet('KB TRACKS')
    ws.cell(row=1, column=1, value=(
        'Progression chains, easiest to hardest. The ballistic chain is the spine of kettlebell '
        'training: the deadlift teaches the swing, the swing teaches the clean and the snatch.')
    ).font = Font(italic=True, size=10)
    ws.merge_cells('A1:D1')
    head(ws, ['Track', 'Rank', 'Movement', 'Level'], row=3)
    byid = {r['id']: r for r in rows}
    j = 4
    for t, members in KB['tracks'].items():
        for i, mid in enumerate(members, 1):
            m = byid[mid]
            for k, v in enumerate([t, i, m['name'], m['level']], start=1):
                c = ws.cell(row=j, column=k, value=v); c.border = THIN
            j += 1
        j += 1
    ws.freeze_panes = 'A4'
    for i, w in enumerate([20, 7, 36, 8], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_complexes(wb):
    ws = wb.create_sheet('KB COMPLEXES')
    ws.cell(row=1, column=1, value=(
        'Named, attributable kettlebell complexes and programmes. These map onto session formats '
        'we already have — ladders, EMOM, clusters — and several are Shocker candidates.')
    ).font = Font(italic=True, size=10)
    ws.merge_cells('A1:F1')
    head(ws, ['Complex', 'Source', 'Use it?', 'Kit', 'Level', 'The protocol'], row=3)
    for j, (name, src, proto, kit, lvl) in enumerate(KB['complexes'], start=4):
        for i, v in enumerate([name, src, '', kit, lvl, proto], start=1):
            c = ws.cell(row=j, column=i, value=v); c.border = THIN
            c.alignment = Alignment(vertical='top', wrap_text=(i == 6))
            if i == 3: c.fill = NEEDS
    last = 3 + len(KB['complexes'])
    d = DataValidation(type='list', formula1='"yes,no,later"', allow_blank=True)
    ws.add_data_validation(d); d.add(f'C4:C{last}')
    ws.freeze_panes = 'A4'
    for i, w in enumerate([28, 22, 10, 10, 8, 88], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_coverage(wb):
    ws = wb.create_sheet('COVERAGE', 0)
    ws.cell(row=1, column=1, value='KETTLEBELL LIBRARY').font = Font(bold=True, size=16)
    ws.cell(row=2, column=1, value=(
        'What bells can and cannot do. Read this before deciding the kettlebell-only promise.')
    ).font = Font(italic=True, size=10)

    r = 4
    ws.cell(row=r, column=1, value='Coverage from this library alone').font = Font(bold=True, size=12); r += 1
    head(ws, ['Pattern group', 'Movements', 'Verdict'], row=r); r += 1
    for g, n in MERGE['after'].items():
        verdict = 'nothing' if n == 0 else ('thin — below the floor of 12' if n < 12 else 'OK')
        for i, v in enumerate([g, n, verdict], start=1):
            c = ws.cell(row=r, column=i, value=v); c.border = THIN
            if i == 3: c.fill = AUTO if verdict == 'OK' else (WARN if n == 0 else NEEDS)
        r += 1

    r += 2
    ws.cell(row=r, column=1, value='Where kettlebells genuinely fail').font = Font(bold=True, size=12); r += 1
    head(ws, ['Pattern', 'The honest position'], row=r); r += 1
    for k, v in KB['coverage_warning'].items():
        for i, val in enumerate([k, v], start=1):
            c = ws.cell(row=r, column=i, value=val); c.border = THIN
            c.alignment = Alignment(vertical='top', wrap_text=(i == 2))
            if i == 1: c.font = Font(bold=True)
            if i == 2 and k == 'v-pull': c.fill = WARN
        ws.row_dimensions[r].height = 58
        r += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 110
    ws.column_dimensions['C'].width = 30


wb = Workbook(); wb.remove(wb.active)
sheet_movements(wb); sheet_tracks(wb); sheet_complexes(wb); sheet_coverage(wb)
wb.active = 0
wb.save(OUT)
print(f'wrote {OUT}')
print(f'  {len(rows)} movements · {len(KB["tracks"])} tracks · {len(KB["complexes"])} complexes')
