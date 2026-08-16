#!/usr/bin/env python3
"""
BUILD THE WORKBOOK

  node "EXERCISE DATABASE/tools/derive.mjs"      # first — writes _derived.json
  python3 "EXERCISE DATABASE/tools/build-workbook.py"

Produces EXERCISE DATABASE/workbook/EXERCISES.xlsx — the authoring surface.
Every enum column gets a dropdown, every sheet gets a filter and a frozen
header, and anything a human still has to decide is coloured.

Re-running REBUILDS from js/data/exercises.js and overwrites the file.
Once Nicolas starts editing, the flow reverses: the workbook becomes the
source and tools/xlsx-to-js.mjs regenerates exercises.js.
"""
import json, os, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

HERE = os.path.dirname(os.path.abspath(__file__))
WB_DIR = os.path.join(HERE, '..', 'workbook')
OUT = os.path.join(WB_DIR, 'EXERCISES.xlsx')

# ── palette ───────────────────────────────────────────────────────────
INK      = '1A1A2E'
HEAD_BG  = '16213E'
HEAD_FG  = 'FFFFFF'
GROUP_BG = {'IDENTITY':'2D3E50','MECHANICS':'1F4E5F','PRESCRIPTION':'4A3A5E',
            'PLACEMENT':'5E4A3A','RELATIONS':'2F5D50','CONSTRAINTS':'6B3A3A',
            'CALIBRATION':'7A5C1E','TEACHING':'3A3A4A','STATUS':'444444',
            'MASTERY':'5A2F52'}
NEEDS    = PatternFill('solid', fgColor='FFF2CC')   # you decide
AUTO     = PatternFill('solid', fgColor='E8F5E9')   # derived, probably fine
BLANK    = PatternFill('solid', fgColor='FFE0E0')   # required + empty
THIN     = Border(*[Side(style='thin', color='D8D8D8')]*4)

# ── closed vocabularies ───────────────────────────────────────────────
ENUMS = {
 'discipline': ['gym','functional','calisthenics'],
 'patterns': ['h-push','v-push','straight-arm-push','h-pull','v-pull','straight-arm-pull',
              'squat','hinge','lunge','calf','shin','carry','anti-extension','anti-rotation',
              'anti-lateral-flexion','flexion','extension','compression','jump','locomotion'],
 'muscles': ['quad','glute','hamstring','adductor','calf','tibialis','hip-flexor','lower-back',
             'lat','mid-back','traps','rear-delt','front-delt','side-delt','chest','upper-chest',
             'biceps','triceps','forearm','grip','abs','obliques','serratus'],
 'laterality': ['bilateral','unilateral','alternating'],
 'measure': ['reps','hold','distance','calories'],
 'loading': ['bodyweight','leverage','added-load','external-load','banded','assisted'],
 'role': ['joint-prep','activation','working-set','skill-strength','skill-practice',
          'finisher','conditioning'],
 'level': ['beg','int','adv'],
 'cns': ['low','moderate','high'],
 'space': ['floor','wall','bar','bench','machine','outdoor'],
 'demands': ['overhead-rom','deep-knee-flexion','wrist-extension','hip-hinge','grip','hang',
             'inversion','impact','floor-to-stand'],
 'contra': ['shoulder','elbow','wrist','neck','lower-back','hip','knee','ankle'],
 'calib_unit': ['reps','reps_per_side','secs','metres','calories'],
 'bool': ['TRUE','FALSE'],
 'fundamental': ['15','30','50'],
 'fund_family': ['squat','hinge','single-leg','h-push','v-push','h-pull','v-pull',
                 'explosive','carry-grip','full-body','core-static','core-dynamic',
                 'mobility','aerobic'],
 'card': ['to write','draft','done'],
 'art': ['missing','source','drawn'],
}

# ── column plan: (key, header, group, width, dropdown, required) ───────
COLS = [
 ('id',               'id',              'IDENTITY',    26, None,          True),
 ('name',             'Name',            'IDENTITY',    32, None,          True),
 ('discipline',       'Discipline',      'IDENTITY',    14, 'discipline',  True),
 ('is_skill',         'Skill?',          'IDENTITY',     8, 'bool',        True),

 ('patterns',         'Patterns',        'MECHANICS',   30, 'patterns',    True),
 ('muscles_primary',  'Muscles · primary','MECHANICS',  26, 'muscles',     True),
 ('muscles_secondary','Muscles · secondary','MECHANICS',28,'muscles',      False),
 ('laterality',       'Laterality',      'MECHANICS',   13, 'laterality',  True),

 ('measure',          'Measure',         'PRESCRIPTION',11, 'measure',     True),
 ('dual',             'Reps OR hold?',   'PRESCRIPTION',13, 'bool',        False),
 ('loading',          'Loading',         'PRESCRIPTION',15, 'loading',     True),
 ('loadable',         'Takes weight?',   'PRESCRIPTION',13, 'bool',        True),
 ('sec_per_rep',      'Sec / rep',       'PRESCRIPTION',10, None,          False),
 ('plyo',             'Plyo?',           'PRESCRIPTION', 8, 'bool',        True),

 ('role',             'Role',            'PLACEMENT',   30, 'role',        True),
 ('level',            'Level',           'PLACEMENT',    8, 'level',       True),
 ('diff',             'Diff 1-10',       'PLACEMENT',   10, None,          True),
 ('cns',              'CNS cost',        'PLACEMENT',   11, 'cns',         True),
 ('ess',              'Essential?',      'PLACEMENT',   11, 'bool',        False),

 ('track',            'Track',           'RELATIONS',   22, None,          False),
 ('rank',             'Rank',            'RELATIONS',    7, None,          False),
 ('families',         'Families',        'RELATIONS',   22, None,          False),
 ('subs',             'Substitutes',     'RELATIONS',   28, None,          False),

 ('equipment',        'Equipment',       'CONSTRAINTS', 22, None,          True),
 ('space',            'Space',           'CONSTRAINTS', 11, 'space',       True),
 ('demands',          'Demands',         'CONSTRAINTS', 32, 'demands',     False),
 ('contra',           'Hard on (joints)','CONSTRAINTS', 26, 'contra',      False),

 ('fundamental',      'Fundamental',     'MASTERY',     13, 'fundamental', False),
 ('fund_no',          '#',               'MASTERY',      8, None,          False),
 ('fund_family',      'Fundamental family','MASTERY',   19, 'fund_family', False),
 ('card',             'Card',            'MASTERY',     11, 'card',        False),
 ('art',              'Artwork',         'MASTERY',     11, 'art',         False),

 ('calib_beg',        'Max · beginner',  'CALIBRATION', 14, None,          False),
 ('calib_int',        'Max · intermediate','CALIBRATION',17,None,          False),
 ('calib_adv',        'Max · advanced',  'CALIBRATION', 15, None,          False),
 ('calib_unit',       'Unit',            'CALIBRATION', 14, 'calib_unit',  False),

 ('no_pr',            'No PR',           'TEACHING',     8, 'bool',        False),
 ('cue',              'Cue',             'TEACHING',    60, None,          False),
 ('standard',         'What counts as a rep','TEACHING', 40, None,         False),

 ('review',           'Needs your eyes on','STATUS',    26, None,          False),
 ('status',           'Status',          'STATUS',      10, None,          False),
]

GYM_COLS = ['Category','Name','Equipment','Pattern','Primary muscles',
            'Already in DB?','Include?','Notes']


def style_header(ws, headers, groups=None):
    """Two header rows when groups are given: group band, then column names."""
    r = 1
    if groups:
        for i, (h, g) in enumerate(zip(headers, groups), start=1):
            c = ws.cell(row=1, column=i, value=g)
            c.fill = PatternFill('solid', fgColor=GROUP_BG.get(g, HEAD_BG))
            c.font = Font(bold=True, color='FFFFFF', size=9)
            c.alignment = Alignment(horizontal='center')
        r = 2
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.fill = PatternFill('solid', fgColor=HEAD_BG)
        c.font = Font(bold=True, color=HEAD_FG, size=10)
        c.alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[r].height = 30
    return r


def build_movements(wb, rows):
    ws = wb.create_sheet('MOVEMENTS')
    headers = [c[1] for c in COLS]
    groups  = [c[2] for c in COLS]
    hr = style_header(ws, headers, groups)

    for j, row in enumerate(rows, start=hr + 1):
        for i, (key, _, _, _, _, req) in enumerate(COLS, start=1):
            v = row.get(key, '')
            cell = ws.cell(row=j, column=i, value=v)
            cell.border = THIN
            cell.alignment = Alignment(vertical='top', wrap_text=(key in ('cue', 'standard')))
            if req and (v == '' or v is None):
                cell.fill = BLANK
        st = ws.cell(row=j, column=len(COLS))
        st.fill = NEEDS if row.get('status') == 'REVIEW' else AUTO
        st.font = Font(bold=True, size=9)

    last = hr + len(rows)
    ws.freeze_panes = f'C{hr+1}'
    ws.auto_filter.ref = f'A{hr}:{get_column_letter(len(COLS))}{last}'
    for i, (_, _, _, w, _, _) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # dropdowns. multi-value columns get a non-blocking hint list so a
    # comma-separated entry is still allowed; single-value columns are strict.
    MULTI = {'patterns', 'muscles', 'role', 'demands', 'contra'}
    for i, (key, hdr, _, _, dv, _) in enumerate(COLS, start=1):
        if not dv:
            continue
        vals = ENUMS[dv]
        strict = dv not in MULTI
        d = DataValidation(type='list', formula1='"%s"' % ','.join(vals),
                           allow_blank=True, showDropDown=False)
        d.error = 'Pick from the list — the build rejects anything else.'
        d.errorTitle = f'{hdr}'
        d.prompt = ('One value.' if strict else
                    'One or more, comma-separated. Allowed: ' + ', '.join(vals))
        d.promptTitle = hdr
        d.showErrorMessage = strict
        d.showInputMessage = True
        ws.add_data_validation(d)
        L = get_column_letter(i)
        d.add(f'{L}{hr+1}:{L}{last}')

    # rows still needing a human, tinted across the whole row
    ws.conditional_formatting.add(
        f'A{hr+1}:{get_column_letter(len(COLS))}{last}',
        FormulaRule(formula=[f'${get_column_letter(len(COLS))}{hr+1}="REVIEW"'],
                    fill=PatternFill('solid', fgColor='FFFBF0'), stopIfTrue=False))
    return ws


def build_tracks(wb, rows):
    """Progression chains, one block per track. The gaps are the work."""
    ws = wb.create_sheet('TRACKS')
    hr = style_header(ws, ['Track', 'Rank', 'Movement', 'Level', 'Diff',
                           'Pattern', 'Discipline', 'Chain status'])
    tracks = {}
    for r in rows:
        if r['track']:
            tracks.setdefault(r['track'], []).append(r)
    untracked = [r for r in rows if not r['track'] and r['role'] != 'joint-prep'
                 and 'joint-prep' not in r['role']]

    j = hr + 1
    for t in sorted(tracks, key=lambda k: -len(tracks[k])):
        members = sorted(tracks[t], key=lambda r: r['rank'])
        n = len(members)
        note = 'OK' if n >= 3 else 'THIN — needs rungs'
        for m in members:
            for i, v in enumerate([t, m['rank'], m['name'], m['level'], m['diff'],
                                   m['patterns'].split(',')[0], m['discipline'], note], start=1):
                c = ws.cell(row=j, column=i, value=v)
                c.border = THIN
                if i == 8 and note != 'OK':
                    c.fill = NEEDS
            j += 1
        j += 1                                  # blank line between chains

    ws.cell(row=j + 1, column=1, value='NO TRACK — these movements sit in no progression chain').font = Font(bold=True, color='B00020')
    j += 2
    for m in sorted(untracked, key=lambda r: (r['discipline'], r['patterns'], r['name'])):
        for i, v in enumerate(['(none)', '', m['name'], m['level'], m['diff'],
                               m['patterns'].split(',')[0], m['discipline'], 'NEEDS A TRACK'], start=1):
            c = ws.cell(row=j, column=i, value=v)
            c.border = THIN
            c.fill = NEEDS
        j += 1

    ws.freeze_panes = f'A{hr+1}'
    for i, w in enumerate([24, 7, 34, 8, 7, 20, 14, 20], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def build_calibration(wb, rows):
    """The batched task. Essentials first — those get prescribed most."""
    ws = wb.create_sheet('CALIBRATION')
    ws.cell(row=1, column=1, value=(
        'How many can a person AT THAT LEVEL do in one honest set today? '
        'Not the rank gate — that lives in BENCHMARKS - Male/Female.csv. '
        'This is what the generator prescribes a percentage of.')).font = Font(italic=True, size=10)
    ws.merge_cells('A1:I1')
    hr = 3
    for i, h in enumerate(['Priority', 'Movement', 'Discipline', 'Pattern', 'Unit',
                           'Beginner', 'Intermediate', 'Advanced', 'Done?'], start=1):
        c = ws.cell(row=hr, column=i, value=h)
        c.fill = PatternFill('solid', fgColor=HEAD_BG)
        c.font = Font(bold=True, color=HEAD_FG)

    def prio(r):
        if r['ess'] == 'TRUE': return (0, 'A · essential')
        if r['level'] == 'beg': return (1, 'B · beginner pool')
        if r['role'].startswith('working-set'): return (2, 'C · working set')
        return (3, 'D · later')

    pool = [r for r in rows if 'working-set' in r['role'] or r['ess'] == 'TRUE']
    pool.sort(key=lambda r: (prio(r)[0], r['discipline'], r['patterns'], r['name']))
    for j, r in enumerate(pool, start=hr + 1):
        for i, v in enumerate([prio(r)[1], r['name'], r['discipline'],
                               r['patterns'].split(',')[0], r['calib_unit'],
                               '', '', '', ''], start=1):
            c = ws.cell(row=j, column=i, value=v)
            c.border = THIN
            if 6 <= i <= 8:
                c.fill = NEEDS
    last = hr + len(pool)
    ws.freeze_panes = f'A{hr+1}'
    ws.auto_filter.ref = f'A{hr}:I{last}'
    for i, w in enumerate([16, 34, 14, 20, 15, 12, 14, 12, 9], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws, len(pool)


def build_fundamentals(wb, rows):
    """The 15 / 30 / 50 a person masters. Cards and artwork attach HERE —
    this is the list the app promises to show and teach, so a gap on it is
    the same class of bug as onboarding asking about a movement we cannot name."""
    ws = wb.create_sheet('FUNDAMENTALS')
    ws.cell(row=1, column=1, value=(
        'Cumulative: advanced = all 50, including the 15. The number is what you have '
        'MASTERED; roughly 12-15 is what you TRAIN in any given week at any level. '
        'Every movement here needs a card and a picture. Source: FOR NICOLAS/FUNDAMENTALS.md')
    ).font = Font(italic=True, size=10)
    ws.merge_cells('A1:H1')
    hr = 3
    for i, h in enumerate(['#','Tier','Movement','Family','In the database?',
                           'Artwork','Card','What is needed'], start=1):
        c = ws.cell(row=hr, column=i, value=h)
        c.fill = PatternFill('solid', fgColor=HEAD_BG)
        c.font = Font(bold=True, color=HEAD_FG)

    with open(os.path.join(HERE, 'fundamentals.json')) as f:
        fund = json.load(f)['list']
    byid = {r['id']: r for r in rows}

    missing = 0
    j = hr + 1
    for no, name, family, tier, exid in fund:
        row = byid.get(exid) if exid else None
        indb = row['name'] if row else 'NO — must be added'
        art  = row['art'] if row else '-'
        card = 'to write'
        need = []
        if not row:  need.append('add the movement')
        elif art != 'drawn': need.append('draw it')
        need.append('write the card')
        if not row: missing += 1
        for i, v in enumerate([no, tier, name, family, indb, art, card, ' + '.join(need)], start=1):
            c = ws.cell(row=j, column=i, value=v)
            c.border = THIN
            if i == 5: c.fill = AUTO if row else BLANK
            if i == 6: c.fill = AUTO if art == 'drawn' else NEEDS
            if i == 7: c.fill = NEEDS
        j += 1
    last = j - 1

    for key, col in (('art','F'), ('card','G')):
        d = DataValidation(type='list', formula1='"%s"' % ','.join(ENUMS[key]), allow_blank=True)
        ws.add_data_validation(d)
        d.add(f'{col}{hr+1}:{col}{last}')

    ws.freeze_panes = f'A{hr+1}'
    ws.auto_filter.ref = f'A{hr}:H{last}'
    for i, w in enumerate([5, 7, 36, 16, 30, 11, 11, 34], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws, len(fund), missing


def build_gym(wb, rows):
    """The parallel classic-gym build. The LIST comes first; the schema
    columns get filled once Nicolas has said yes / no / add to each line."""
    ws = wb.create_sheet('GYM LIBRARY')
    ws.cell(row=1, column=1, value=(
        'PARALLEL TRACK — the classic gym library. Say yes / no / add on each line first; '
        'we fill the full schema for the keepers afterwards. "Already in DB?" is checked '
        'automatically against js/data/exercises.js.')).font = Font(italic=True, size=10)
    ws.merge_cells('A1:H1')
    hr = 3
    for i, h in enumerate(GYM_COLS, start=1):
        c = ws.cell(row=hr, column=i, value=h)
        c.fill = PatternFill('solid', fgColor=HEAD_BG)
        c.font = Font(bold=True, color=HEAD_FG)

    with open(os.path.join(HERE, 'gym-library-seed.json')) as f:
        seed = json.load(f)['movements']

    def norm(s):
        return re.sub(r'[^a-z]', '', s.lower())
    existing = {norm(r['name']): r['name'] for r in rows}

    j = hr + 1
    for cat, name, equip, pat, mus in seed:
        hit = existing.get(norm(name), '')
        for i, v in enumerate([cat, name, equip, pat, mus,
                               hit or 'no', '', ''], start=1):
            c = ws.cell(row=j, column=i, value=v)
            c.border = THIN
            if i == 6:
                c.fill = AUTO if hit else NEEDS
            if i == 7:
                c.fill = NEEDS
        j += 1
    last = j - 1

    d = DataValidation(type='list', formula1='"yes,no,later"', allow_blank=True)
    ws.add_data_validation(d)
    d.add(f'G{hr+1}:G{last}')

    ws.freeze_panes = f'A{hr+1}'
    ws.auto_filter.ref = f'A{hr}:H{last}'
    for i, w in enumerate([16, 34, 22, 20, 28, 16, 11, 34], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws, len(seed), sum(1 for cat, n, *_ in seed if norm(n) in existing)


def build_enums(wb):
    ws = wb.create_sheet('LISTS')
    ws.cell(row=1, column=1, value=(
        'The allowed values. Anything outside these fails the build — that is what '
        'stops a fourteenth spelling of "push" appearing next month.')).font = Font(italic=True, size=10)
    ws.merge_cells('A1:H1')
    col = 1
    for key, vals in ENUMS.items():
        c = ws.cell(row=3, column=col, value=key)
        c.fill = PatternFill('solid', fgColor=HEAD_BG)
        c.font = Font(bold=True, color=HEAD_FG)
        for k, v in enumerate(vals, start=4):
            ws.cell(row=k, column=col, value=v).border = THIN
        ws.column_dimensions[get_column_letter(col)].width = max(14, len(key) + 4)
        col += 1
    ws.freeze_panes = 'A4'
    return ws


def build_status(wb, rows, ncal, ngym, ngymhave, nfund, nfundmiss):
    ws = wb.create_sheet('START HERE', 0)
    W = [('A', 40), ('B', 16), ('C', 16), ('D', 16), ('E', 16), ('F', 40)]
    for L, w in W:
        ws.column_dimensions[L].width = w

    def title(r, text, size=16):
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(bold=True, size=size, color=INK)

    def head(r, cells):
        for i, v in enumerate(cells, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.fill = PatternFill('solid', fgColor=HEAD_BG)
            c.font = Font(bold=True, color=HEAD_FG)

    title(1, 'EXERCISE DATABASE — start here')
    ws.cell(row=2, column=1, value='Rebuilt by tools/build-workbook.py. Sheets: MOVEMENTS · TRACKS · CALIBRATION · GYM LIBRARY · LISTS').font = Font(italic=True, size=10)

    # curation state by discipline x pattern group
    title(4, 'Where curation stands', 13)
    head(5, ['Area', 'Movements', 'Have level+diff', 'In a track', 'Calibrated', 'State'])
    GROUPS = [
        ('Core',       lambda r: r['patterns'].split(',')[0].strip() in
                       ('anti-extension','anti-rotation','anti-lateral-flexion','flexion','extension','compression')),
        ('Push',       lambda r: r['patterns'].split(',')[0].strip() in ('h-push','v-push','straight-arm-push')),
        ('Pull',       lambda r: r['patterns'].split(',')[0].strip() in ('h-pull','v-pull','straight-arm-pull')),
        ('Lower body', lambda r: r['patterns'].split(',')[0].strip() in ('squat','hinge','lunge','calf','shin')),
        ('Full body / conditioning', lambda r: r['patterns'].split(',')[0].strip() in ('jump','locomotion','carry')),
        ('Mobility / prep', lambda r: not r['patterns'].strip()),
    ]
    STATE = {'Core':'DONE','Push':'DONE','Pull':'TO DO','Lower body':'TO DO',
             'Full body / conditioning':'TO DO','Mobility / prep':'TO DO'}
    r = 6
    sizes = {}
    for label, pred in GROUPS:
        sub = [x for x in rows if pred(x)]
        sizes[label] = len(sub)
        lvl = sum(1 for x in sub if x['level'] and x['diff'] != '')
        trk = sum(1 for x in sub if x['track'])
        for i, v in enumerate([label, len(sub), f'{lvl} / {len(sub)}',
                               f'{trk} / {len(sub)}', f'0 / {len(sub)}', STATE[label]], start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = THIN
            if i == 6:
                c.fill = AUTO if v == 'DONE' else NEEDS
                c.font = Font(bold=True)
        r += 1

    # what a person can actually be given
    r += 2
    title(r, 'What the database can offer a real person', 13); r += 1
    head(r, ['Profile', 'Reachable', 'Push', 'Pull', 'Lower', 'Verdict']); r += 1
    PROFILES = [
        ('Ozzy — kettlebell, bench, mat', {'bw','kb','bench','mat'}),
        ('Bodyweight only', {'bw','mat'}),
        ('Home + dumbbells', {'bw','mat','db','bench','band'}),
        ('Full gym', None),
    ]
    def kitok(row, kit):
        if kit is None: return True
        return all(e.strip() in kit for e in row['equipment'].split(',') if e.strip())
    for label, kit in PROFILES:
        sub = [x for x in rows if kitok(x, kit)]
        cnt = lambda ps: sum(1 for x in sub if x['patterns'].split(',')[0].strip() in ps)
        push = cnt({'h-push','v-push','straight-arm-push'})
        pull = cnt({'h-pull','v-pull','straight-arm-pull'})
        low  = cnt({'squat','hinge','lunge','calf','shin'})
        verdict = 'OK' if min(push, pull, low) >= 12 else f'THIN — {min(push,pull,low)} in the smallest bucket'
        for i, v in enumerate([label, len(sub), push, pull, low, verdict], start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = THIN
            if i == 6:
                c.fill = AUTO if verdict == 'OK' else NEEDS
                c.font = Font(bold=True)
        r += 1

    # the tasks
    r += 2
    title(r, 'What is yours to do', 13); r += 1
    head(r, ['Task', 'Sheet', 'Size', '', '', 'Why it matters']); r += 1
    TASKS = [
      ('Say yes / no on the gym list', 'GYM LIBRARY', f'{ngym} lines', '', '',
       f'{ngymhave} already exist; the rest is the parallel build'),
      ('Fill Beg / Int / Adv max reps', 'CALIBRATION', f'{ncal} rows, batch ~25', '', '',
       'The one thing nobody but you can supply — every prescription reads it'),
      ('Curate Pull', 'MOVEMENTS · filter Pattern', f'{sizes["Pull"]} rows', '', '',
       'Core and Push are done; Pull is next'),
      ('Curate Lower body', 'MOVEMENTS · filter Pattern', f'{sizes["Lower body"]} rows', '', '',
       'Never curated. 2 hamstring movements in the whole database'),
      ('Curate Full body / conditioning', 'MOVEMENTS · filter Pattern',
       f'{sizes["Full body / conditioning"]} rows', '', '',
       'Almost nothing there yet'),
      ('Check the yellow rows', 'MOVEMENTS · filter Status', f"{sum(1 for x in rows if x['status']=='REVIEW')} rows", '', '',
       'Where I guessed and want your eyes'),
      ('Add the missing fundamentals', 'FUNDAMENTALS', f'{nfundmiss} of {nfund} absent', '', '',
       'Movements we promise to teach to mastery that the database does not hold'),
      ('Write the exercise cards', 'FUNDAMENTALS', f'{nfund} cards', '', '',
       'One per fundamental. The card is what the app shows when you tap the movement'),
      ('Finish the artwork', 'MOVEMENTS · filter Artwork',
       f"{sum(1 for x in rows if x['art']!='drawn')} without a picture", '', '',
       f"{sum(1 for x in rows if x['fundamental'] and x['art']!='drawn')} of those are fundamentals — do those first"),
    ]
    for t in TASKS:
        for i, v in enumerate(t, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = THIN
            if i == 1:
                c.font = Font(bold=True)
        r += 1

    r += 2
    ws.cell(row=r, column=1, value='Colour key').font = Font(bold=True, size=12); r += 1
    for label, fill in [('You decide / I guessed', NEEDS),
                        ('Derived, probably fine', AUTO),
                        ('Required and empty', BLANK)]:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2).fill = fill
        ws.cell(row=r, column=2).border = THIN
        r += 1
    return ws


def main():
    with open(os.path.join(WB_DIR, '_derived.json')) as f:
        rows = json.load(f)

    wb = Workbook()
    wb.remove(wb.active)
    build_movements(wb, rows)
    build_tracks(wb, rows)
    _, ncal = build_calibration(wb, rows)
    _, nfund, nfundmiss = build_fundamentals(wb, rows)
    _, ngym, ngymhave = build_gym(wb, rows)
    build_enums(wb)
    build_status(wb, rows, ncal, ngym, ngymhave, nfund, nfundmiss)
    wb.active = 0
    wb.save(OUT)

    print(f'wrote {OUT}')
    print(f'  MOVEMENTS    {len(rows)} rows, {sum(1 for r in rows if r["status"]=="REVIEW")} flagged for review')
    print(f'  CALIBRATION  {ncal} rows to fill')
    print(f'  GYM LIBRARY  {ngym} proposed, {ngymhave} already in the database')


if __name__ == '__main__':
    main()
