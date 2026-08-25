#!/usr/bin/env python3
"""
EXTRACT DECISIONS — pull Nicolas's answers out of the workbooks into JSON.

  python3 "EXERCISE DATABASE/tools/extract-decisions.py"

Everything he types into a sheet lands in workbook/_decisions.json, which the
build reads. That is what makes the rebuild safe: his judgement lives in a file
the generator never writes, so regenerating the workbook can no longer lose it.

Run this BEFORE any rebuild, every time.
"""
import json, os
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
WB   = os.path.join(HERE, '..', 'workbook')
OUT  = os.path.join(WB, '_decisions.json')

def txt(c):
    v = c.value
    return v.strip().lower() if isinstance(v, str) else ('' if v is None else str(v).strip().lower())

def raw(c):
    v = c.value
    return v.strip() if isinstance(v, str) else v

# Start from whatever is already recorded and MERGE. A sheet that has been
# folded into the database no longer exists to be re-read, and a fresh dict
# would silently wipe every answer that came from it.
try:
    with open(OUT) as _f:
        d = json.load(_f)
except Exception:
    d = {}
for _k in ('gym', 'gym_notes', 'kettlebell', 'complexes', 'calibration',
           'cards', 'movement_edits', 'nick_notes'):
    d.setdefault(_k, {})

# A stray copy — "EXERCISES_after_crash.xlsx", "EXERCISES (1).xlsx", a Downloads
# copy — is how a whole afternoon of answers goes missing. Excel makes these
# without asking. Refuse to run quietly when one exists.
import glob
strays = [f for f in glob.glob(os.path.join(WB, '*.xlsx'))
          if os.path.basename(f) != 'EXERCISES.xlsx' and not os.path.basename(f).startswith('~$')]
if strays:
    print('STOP — workbook/ should hold EXERCISES.xlsx and nothing else:')
    for f in strays:
        print('   ', os.path.basename(f))
    print('    Only EXERCISES.xlsx is read. Merge the stray in (or delete it) before continuing,')
    print('    otherwise the answers in it are invisible to the build.')
    raise SystemExit(1)

# ---- EXERCISES.xlsx -------------------------------------------------
p = os.path.join(WB, 'EXERCISES.xlsx')
if os.path.exists(p):
    wb = load_workbook(p)

    ws = wb['GYM LIBRARY']
    for r in range(4, ws.max_row + 1):
        name = raw(ws.cell(r, 2))
        if not name: continue
        ans = txt(ws.cell(r, 7))
        if ans: d['gym'][name] = ans
        note = raw(ws.cell(r, 8))
        if note: d['gym_notes'][name] = note

    # CALIBRATION now ships PRE-FILLED with my proposed numbers, so "cell has a
    # value" no longer means "Nicolas decided this". The Source column is the
    # discriminator: only 'yours' is his. Anything he overtypes on a proposed
    # row is caught too, by comparing against what was proposed.
    try:
        _prop = json.load(open(os.path.join(HERE, 'calibration-proposed.json')))['values']
    except Exception:
        _prop = {}

    ws = wb['CALIBRATION']
    hdr = [ws.cell(3, i).value for i in range(1, ws.max_column + 1)]
    SRC = hdr.index('Source') + 1 if 'Source' in hdr else None
    for r in range(4, ws.max_row + 1):
        name = raw(ws.cell(r, 2))
        if not name: continue
        vals = {k: raw(ws.cell(r, i)) for k, i in (('beg', 6), ('int', 7), ('adv', 8))}
        wts  = {k: raw(ws.cell(r, i)) for k, i in (('beg', 9), ('int', 10), ('adv', 11))}
        if not any(v not in (None, '') for v in vals.values()):
            continue
        src = txt(ws.cell(r, SRC)) if SRC else ''
        if src == 'yours':
            d['calibration'][name] = vals
        elif src == 'proposed':
            # only record it if he changed one of the three numbers
            p3 = _prop.get(name)
            if p3 and [vals['beg'], vals['int'], vals['adv']] != [p3[0], p3[1], p3[2]]:
                d['calibration'][name] = vals
        else:
            d['calibration'][name] = vals
        if any(v not in (None, '') for v in wts.values()):
            d.setdefault('calibration_weights', {})[name] = wts

    # The ladder layout puts the fundamental's name in column 1 and the card
    # in the last column. Reading the old positions grabbed tier numbers as
    # names and rung names as card text.
    ws = wb['FUNDAMENTALS']
    hdr = [ws.cell(3, i).value for i in range(1, ws.max_column + 1)]
    CARD = hdr.index('Card') + 1 if 'Card' in hdr else None
    if CARD:
        for r in range(4, ws.max_row + 1):
            name = raw(ws.cell(r, 1))
            if not name or str(name).startswith('   '): continue
            card = raw(ws.cell(r, CARD))
            if card and str(card).strip().lower() != 'to write':
                d['cards'][name] = card

# ---- KETTLEBELL.xlsx ------------------------------------------------
# KETTLEBELL.xlsx was folded into EXERCISES.xlsx on 20 Aug. If it is still
# around it is read; if not, the answers already in _decisions.json stand.
p = os.path.join(WB, 'KETTLEBELL.xlsx')
if os.path.exists(p):
    wb = load_workbook(p)
    ws = wb['KETTLEBELL']
    for r in range(4, ws.max_row + 1):
        mid = raw(ws.cell(r, 1))
        if not mid: continue
        ans = txt(ws.cell(r, 3))
        if ans: d['kettlebell'][mid] = ans
    ws = wb['KB COMPLEXES']
    for r in range(4, ws.max_row + 1):
        name = raw(ws.cell(r, 1))
        if not name: continue
        ans = txt(ws.cell(r, 3))
        if ans: d['complexes'][name] = ans

with open(OUT, 'w') as f:
    json.dump(d, f, indent=1)

print(f'wrote {OUT}')
for k in ('gym', 'kettlebell', 'complexes', 'calibration', 'cards'):
    n = len(d[k])
    if k == 'gym' and n:
        import collections
        print(f'  {k:12s} {n:4d}  {dict(collections.Counter(d[k].values()))}')
    else:
        print(f'  {k:12s} {n:4d}')
if d['gym_notes']:
    print('  notes:')
    for k, v in d['gym_notes'].items(): print(f'    {k}: {v}')
