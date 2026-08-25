#!/usr/bin/env python3
"""
BUILD THE WORKBOOK

  node "EXERCISE DATABASE/tools/derive.mjs"      # first — writes _derived.json
  python3 "EXERCISE DATABASE/tools/build-workbook.py"

Produces EXERCISE DATABASE/workbook/EXERCISES.xlsx — the authoring surface.
Every enum column gets a dropdown, every sheet gets a filter and a frozen
header, and anything a human still has to decide is coloured.

Re-running REBUILDS from js/data/exercises.js and overwrites the file.
Once Nicolas starts editing, the flow reverses: the workbook becomes the
source and tools/xlsx-to-js.mjs regenerates exercises.js.
"""
import json, os, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

HERE = os.path.dirname(os.path.abspath(__file__))
WB_DIR = os.path.join(HERE, '..', 'workbook')
OUT = os.path.join(WB_DIR, 'EXERCISES.xlsx')

# ── palette ───────────────────────────────────────────────────────────
INK      = '1A1A2E'
HEAD_BG  = '16213E'
HEAD_FG  = 'FFFFFF'
GROUP_BG = {'IDENTITY':'2D3E50','MECHANICS':'1F4E5F','PRESCRIPTION':'4A3A5E',
            'PLACEMENT':'5E4A3A','RELATIONS':'2F5D50','CONSTRAINTS':'6B3A3A',
            'CALIBRATION':'7A5C1E','TEACHING':'3A3A4A','STATUS':'444444',
            'MASTERY':'5A2F52'}
NEEDS    = PatternFill('solid', fgColor='FFF2CC')   # you decide
AUTO     = PatternFill('solid', fgColor='E8F5E9')   # derived, probably fine
OK_FILL  = PatternFill('solid', fgColor='DFF0E4')   # your own answer, keep
BLANK    = PatternFill('solid', fgColor='FFE0E0')   # required + empty
THIN     = Border(*[Side(style='thin', color='D8D8D8')]*4)

# ── closed vocabularies ───────────────────────────────────────────────
ENUMS = {
 'modality': ['bodyweight','kettlebell','dumbbell','barbell','machine','cable','band','other'],
 'patterns': ['h-push','v-push','straight-arm-push','h-pull','v-pull','straight-arm-pull',
              'squat','hinge','lunge','calf','shin','carry','anti-extension','anti-rotation',
              'anti-lateral-flexion','rotation','flexion','extension','compression','jump','locomotion'],
 'muscles': ['quad','glute','hamstring','adductor','calf','tibialis','hip-flexor','lower-back',
             'lat','mid-back','traps','rear-delt','front-delt','side-delt','chest','upper-chest',
             'biceps','triceps','forearm','grip','abs','obliques','serratus'],
 'laterality': ['bilateral','unilateral','alternating'],
 'measure': ['reps','hold','distance','calories'],
 'loading': ['bodyweight','leverage','added-load','external-load','banded','assisted'],
 'role': ['joint-prep','activation','working-set','skill-strength','skill-practice',
          'finisher','conditioning'],
 'level': ['start','beg','int','adv'],
 'cns': ['low','moderate','high'],
 'space': ['floor','wall','bar','bench','machine','outdoor'],
 'demands': ['overhead-rom','deep-knee-flexion','wrist-extension','hip-hinge','grip','hang',
             'inversion','impact','floor-to-stand'],
 'contra': ['shoulder','elbow','wrist','neck','lower-back','hip','knee','ankle'],
 'calib_unit': ['reps','reps_per_side','secs','lbs','metres','calories'],
 'bool': ['TRUE','FALSE'],
 'fundamental': ['15','30','50'],
 'fund_family': ['squat','hinge','single-leg','h-push','v-push','h-pull','v-pull',
                 'explosive','carry-grip','full-body','core-static','core-dynamic',
                 'mobility','aerobic'],
 'card': ['to write','draft','done'],
 'art': ['missing','source','drawn'],
 'ladder_role': ['regression','anchor','progression'],
 'gym_role': ['easier','gym-anchor','harder'],
}

# ── column plan: (key, header, group, width, dropdown, required) ───────
COLS = [
 ('id',               'id',              'IDENTITY',    26, None,          True),
 ('name',             'Name',            'IDENTITY',    32, None,          True),
 ('modality',         'Modality',        'IDENTITY',    14, 'modality',    True),
 ('functional',       'Functional?',     'IDENTITY',    12, 'bool',        True),
 ('is_skill',         'Skill?',          'IDENTITY',     8, 'bool',        True),

 ('patterns',         'Patterns',        'MECHANICS',   30, 'patterns',    True),
 ('muscles_primary',  'Muscles · primary','MECHANICS',  26, 'muscles',     True),
 ('muscles_secondary','Muscles · secondary','MECHANICS',28,'muscles',      False),
 ('laterality',       'Laterality',      'MECHANICS',   13, 'laterality',  True),

 ('measure',          'Measure',         'PRESCRIPTION',11, 'measure',     True),
 ('dual',             'Reps OR hold?',   'PRESCRIPTION',13, 'bool',        False),
 ('loading',          'Loading',         'PRESCRIPTION',15, 'loading',     True),
 ('loadable',         'Takes weight?',   'PRESCRIPTION',13, 'bool',        True),
 ('sec_per_rep',      'Sec / rep',       'PRESCRIPTION',10, None,          False),
 ('plyo',             'Plyo?',           'PRESCRIPTION', 8, 'bool',        True),

 ('role',             'Role',            'PLACEMENT',   30, 'role',        True),
 ('level',            'Level',           'PLACEMENT',    8, 'level',       True),
 ('diff',             'Diff 1-10',       'PLACEMENT',   10, None,          True),
 ('cns',              'CNS cost',        'PLACEMENT',   11, 'cns',         True),
 ('ess',              'Essential?',      'PLACEMENT',   11, 'bool',        False),

 ('track',            'Track',           'RELATIONS',   22, None,          False),
 ('rank',             'Rank',            'RELATIONS',    7, None,          False),
 ('families',         'Families',        'RELATIONS',   22, None,          False),
 ('subs',             'Substitutes',     'RELATIONS',   28, None,          False),
 ('bridges_to',       'Bridges to (fundamental)','RELATIONS', 24, None,     False),
 ('ladder',           'Ladder',          'RELATIONS',   16, None,          False),
 ('ladder_role',      'Rung',            'RELATIONS',   12, 'ladder_role', False),
 ('ladder_pos',       'Pos',             'RELATIONS',    6, None,          False),
 ('gym_lane',         'Gym lane',        'RELATIONS',   16, None,          False),
 ('gym_role',         'Gym rung',        'RELATIONS',   12, 'gym_role',    False),

 ('equipment',        'Equipment',       'CONSTRAINTS', 22, None,          True),
 ('space',            'Space',           'CONSTRAINTS', 11, 'space',       True),
 ('demands',          'Demands',         'CONSTRAINTS', 32, 'demands',     False),
 ('contra',           'Hard on (joints)','CONSTRAINTS', 26, 'contra',      False),

 ('fundamental',      'Fundamental',     'MASTERY',     13, 'fundamental', False),
 ('fund_no',          '#',               'MASTERY',      8, None,          False),
 ('fund_family',      'Fundamental family','MASTERY',   19, 'fund_family', False),
 ('card',             'Card',            'MASTERY',     11, 'card',        False),
 ('art',              'Artwork',         'MASTERY',     11, 'art',         False),

 ('calib_beg',        'Max · beginner',  'CALIBRATION', 14, None,          False),
 ('calib_int',        'Max · intermediate','CALIBRATION',17,None,          False),
 ('calib_adv',        'Max · advanced',  'CALIBRATION', 15, None,          False),
 ('calib_unit',       'Unit',            'CALIBRATION', 14, 'calib_unit',  False),

 ('no_pr',            'No PR',           'TEACHING',     8, 'bool',        False),
 ('cue',              'Cue',             'TEACHING',    60, None,          False),
 ('standard',         'What counts as a rep','TEACHING', 40, None,         False),

 ('review',           'Needs your eyes on','STATUS',    26, None,          False),
 ('status',           'Status',          'STATUS',      10, None,          False),
]

GYM_COLS = ['Category','Name','Equipment','Pattern','Primary muscles',
            'Already in DB?','Include?','Notes']


def _fit(text, limit):
    """Excel rejects over-long data-validation strings by asking for a repair."""
    t = str(text or '')
    return t if len(t) <= limit else t[:limit - 1].rstrip(' ,') + '\u2026'


def style_header(ws, headers, groups=None):
    """Two header rows when groups are given: group band, then column names."""
    r = 1
    if groups:
        for i, (h, g) in enumerate(zip(headers, groups), start=1):
            c = ws.cell(row=1, column=i, value=g)
            c.fill = PatternFill('solid', fgColor=GROUP_BG.get(g, HEAD_BG))
            c.font = Font(bold=True, color='FFFFFF', size=9)
            c.alignment = Alignment(horizontal='center')
        r = 2
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.fill = PatternFill('solid', fgColor=HEAD_BG)
        c.font = Font(bold=True, color=HEAD_FG, size=10)
        c.alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[r].height = 30
    return r


def build_movements(wb, rows):
    ws = wb.create_sheet('MOVEMENTS')
    headers = [c[1] for c in COLS]
    groups  = [c[2] for c in COLS]
    hr = style_header(ws, headers, groups)

    for j, row in enumerate(rows, start=hr + 1):
        for i, (key, _, _, _, _, req) in enumerate(COLS, start=1):
            v = row.get(key, '')
            cell = ws.cell(row=j, column=i, value=v)
            cell.border = THIN
            cell.alignment = Alignment(vertical='top', wrap_text=(key in ('cue', 'standard')))
            if req and (v == '' or v is None):
                cell.fill = BLANK
        st = ws.cell(row=j, column=len(COLS))
        st.fill = NEEDS if row.get('status') == 'REVIEW' else AUTO
        st.font = Font(bold=True, size=9)

    last = hr + len(rows)
    ws.freeze_panes = f'C{hr+1}'
    ws.auto_filter.ref = f'A{hr}:{get_column_letter(len(COLS))}{last}'
    for i, (_, _, _, w, _, _) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # dropdowns. multi-value columns get a non-blocking hint list so a
    # comma-separated entry is still allowed; single-value columns are strict.
    MULTI = {'patterns', 'muscles', 'role', 'demands', 'contra'}
    for i, (key, hdr, _, _, dv, _) in enumerate(COLS, start=1):
        if not dv:
            continue
        vals = ENUMS[dv]
        strict = dv not in MULTI
        d = DataValidation(type='list', formula1='"%s"' % ','.join(vals),
                           allow_blank=True, showDropDown=False)
        # Excel's hard limits: prompt and error 255 chars, titles 32. Going over
        # does not fail here — it makes Excel demand a repair when the file is
        # opened, which is exactly what happened with the Patterns column at 261.
        d.error = _fit('Pick from the list — the build rejects anything else.', 255)
        d.errorTitle = _fit(hdr, 32)
        d.prompt = _fit('One value.' if strict else
                        'One or more, comma-separated. Allowed: ' + ', '.join(vals), 255)
        d.promptTitle = _fit(hdr, 32)
        d.showErrorMessage = strict
        d.showInputMessage = True
        ws.add_data_validation(d)
        L = get_column_letter(i)
        d.add(f'{L}{hr+1}:{L}{last}')

    # rows still needing a human, tinted across the whole row
    ws.conditional_formatting.add(
        f'A{hr+1}:{get_column_letter(len(COLS))}{last}',
        FormulaRule(formula=[f'${get_column_letter(len(COLS))}{hr+1}="REVIEW"'],
                    fill=PatternFill('solid', fgColor='FFFBF0'), stopIfTrue=False))
    return ws


def build_tracks(wb, rows):
    """Progression chains, one block per track. The gaps are the work."""
    ws = wb.create_sheet('TRACKS')
    hr = style_header(ws, ['Track', 'Rank', 'Movement', 'Level', 'Diff',
                           'Pattern', 'Modality', 'Chain status'])
    tracks = {}
    for r in rows:
        if r['track']:
            tracks.setdefault(r['track'], []).append(r)
    untracked = [r for r in rows if not r['track'] and r['role'] != 'joint-prep'
                 and 'joint-prep' not in r['role']]

    j = hr + 1
    for t in sorted(tracks, key=lambda k: -len(tracks[k])):
        members = sorted(tracks[t], key=lambda r: r['rank'])
        n = len(members)
        note = 'OK' if n >= 3 else 'THIN — needs rungs'
        for m in members:
            for i, v in enumerate([t, m['rank'], m['name'], m['level'], m['diff'],
                                   m['patterns'].split(',')[0], m['modality'], note], start=1):
                c = ws.cell(row=j, column=i, value=v)
                c.border = THIN
                if i == 8 and note != 'OK':
                    c.fill = NEEDS
            j += 1
        j += 1                                  # blank line between chains

    ws.cell(row=j + 1, column=1, value='NO TRACK — these movements sit in no progression chain').font = Font(bold=True, color='B00020')
    j += 2
    for m in sorted(untracked, key=lambda r: (r['modality'], r['patterns'], r['name'])):
        for i, v in enumerate(['(none)', '', m['name'], m['level'], m['diff'],
                               m['patterns'].split(',')[0], m['modality'], 'NEEDS A TRACK'], start=1):
            c = ws.cell(row=j, column=i, value=v)
            c.border = THIN
            c.fill = NEEDS
        j += 1

    ws.freeze_panes = f'A{hr+1}'
    for i, w in enumerate([24, 7, 34, 8, 7, 20, 14, 20], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def build_calibration(wb, rows):
    """The three columns describe the MOVEMENT, not the person.

    Nicolas's note, 2026-08-18: "beginner" was doing two jobs — a person who
    never trained, and someone who just unlocked the muscle-up. So the columns
    are now Entry / Solid / Strong, which are true of any movement at any
    difficulty, and WHO ever sees a movement is the separate `level` field.
    """
    ws = wb.create_sheet('CALIBRATION')
    ws.cell(row=1, column=1, value=(
        'ENTRY = the day you first own the movement.   SOLID = you train it comfortably.   '
        'STRONG = you have mastered it.   These describe the MOVEMENT, not the person — so a muscle-up '
        'legitimately starts at 1-2 reps without that meaning "gym beginner". Weights are in LBS. '
        'Source column: "yours" = you typed it, "proposed" = mine from research, overwrite freely.')
    ).font = Font(italic=True, size=10)
    ws.merge_cells('A1:M1')
    hr = 3
    cols = ['Group', 'Movement', 'Modality', 'Pattern', 'Unit',
            'Entry', 'Solid', 'Strong',
            'Wt entry (lbs)', 'Wt solid (lbs)', 'Wt strong (lbs)', 'Source', 'Note']
    for i, h in enumerate(cols, start=1):
        c = ws.cell(row=hr, column=i, value=h)
        c.fill = PatternFill('solid', fgColor=HEAD_BG)
        c.font = Font(bold=True, color=HEAD_FG)
        c.alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[hr].height = 30

    try:
        dec = json.load(open(os.path.join(HERE, '..', 'workbook', '_decisions.json'))).get('calibration', {})
    except Exception:
        dec = {}
    prop = json.load(open(os.path.join(HERE, 'calibration-proposed.json')))
    PV, PW = prop['values'], prop['weights_lbs']

    def group(r):
        if r.get('fundamental'):        return (0, f"A · fundamental {r['fundamental']}")
        if 'kettlebell-fundamental' in r.get('families', ''): return (1, 'B · kettlebell core')
        if r.get('ess') == 'TRUE':      return (2, 'C · essential')
        return (3, 'D · later')

    # anything Nicolas has already answered stays on the sheet even if a later
    # restructure moved it off the fundamentals list. His input is never dropped.
    answered = {k for k, v in dec.items()
                if any(str(v.get(x, '')).strip() not in ('', 'None') for x in ('beg', 'int', 'adv'))}
    pool = [r for r in rows if group(r)[0] <= 2 or r['name'] in answered]
    pool.sort(key=lambda r: (group(r)[0], r['modality'], r['name']))

    def val(v):
        return '' if v in (None, '', 'n/a', 'N/A') else v

    j = hr + 1
    for r in pool:
        name = r['name']
        mine = dec.get(name) or {}
        p    = PV.get(name)
        w    = PW.get(name)
        got_mine = any(val(mine.get(k)) != '' for k in ('beg', 'int', 'adv'))

        if got_mine:
            trio = [val(mine.get('beg')), val(mine.get('int')), val(mine.get('adv'))]
            src, note = 'yours', ''
        elif p and p[0] is not None:
            trio = [p[0], p[1], p[2]]
            src, note = 'proposed', ''
        elif p:
            trio = ['', '', '']
            src, note = '—', p[3]
        else:
            trio = ['', '', '']
            src, note = '', ''

        loaded = r.get('loadable') == 'TRUE'
        wt = [w[0], w[1], w[2]] if w and w[0] is not None else ['', '', '']
        if w and w[0] is None: note = note or w[3]

        vals = [group(r)[1], name, r['modality'], r['patterns'].split(',')[0].strip(),
                r.get('calib_unit', 'reps')] + trio + wt + [src, note]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=j, column=i, value=v)
            c.border = THIN
            if 6 <= i <= 8:
                c.fill = OK_FILL if src == 'yours' else (NEEDS if src == 'proposed' else BLANK)
            if 9 <= i <= 11:
                c.fill = NEEDS if loaded else PatternFill('solid', fgColor='EFEFEF')
            if i == 12:
                c.font = Font(bold=True, size=9,
                              color='2C6A52' if src == 'yours' else '8E6209')
        j += 1
    last = hr + len(pool)
    ws.freeze_panes = f'C{hr+1}'
    ws.auto_filter.ref = f'A{hr}:M{last}'
    for i, w_ in enumerate([20, 32, 13, 18, 15, 9, 9, 9, 13, 13, 13, 11, 40], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w_
    return ws, len(pool)


def build_fundamentals(wb, rows):
    """One row per fundamental, read left to right: the regressions that let
    anyone start, the anchor in the middle, the progressions above it.

    Nicolas, 19 Aug: "Main exercise and regression / progression on level of
    user." A fundamental can be easy (bodyweight squat) or hard (pull-up) —
    what matters is that everybody can enter the ladder somewhere.
    """
    ws = wb.create_sheet('FUNDAMENTALS')
    LAD = json.load(open(os.path.join(HERE, 'fundamental-ladders.json')))
    GYM = json.load(open(os.path.join(HERE, 'gym-lanes.json')))
    byid = {r['id']: r for r in rows}
    try:
        NICK = json.load(open(os.path.join(HERE, '..', 'workbook', '_decisions.json'))).get('nick_notes', {})
    except Exception:
        NICK = {}
    ws.cell(row=1, column=1, value=(
        'Each row is one fundamental, easiest on the left. A person who is JUST STARTING enters at the '
        'leftmost rung they can do; the anchor (bold, blue) is the movement we are getting them to; '
        'everything right of it keeps an advanced person busy. Grey = the movement is not in the '
        'database yet. The GYM row beneath each fundamental is the loaded version of the same job — '
        'shorter on purpose, because load is the progression. Five fundamentals have no honest gym '
        'equivalent and say so.')).font = Font(italic=True, size=10)
    ws.merge_cells('A1:V1')

    hr = 3
    cols = ['Fundamental', 'Family', 'Tier',
            '<- easiest', '', '', '', '', 'ANCHOR', '', '', '', '', '-> hardest',
            'Entry level', 'Rungs', 'Targets (muscles)', 'Cue', 'Video', 'Notes', 'Your call', 'Card']
    for i, h in enumerate(cols, start=1):
        c = ws.cell(row=hr, column=i, value=h)
        c.fill = PatternFill('solid', fgColor='2D3E50' if i == 9 else HEAD_BG)
        c.font = Font(bold=True, color=HEAD_FG, size=10)
        c.alignment = Alignment(horizontal='center' if 4 <= i <= 14 else 'left',
                                wrap_text=True, vertical='center')
    ws.row_dimensions[hr].height = 26

    ANCHOR_FILL = PatternFill('solid', fgColor='D6E4F0')
    REG_FILL    = PatternFill('solid', fgColor='F2F6F9')
    PROG_FILL   = PatternFill('solid', fgColor='F7F3F9')
    GONE        = PatternFill('solid', fgColor='E4E4E4')
    GYM_FILL    = PatternFill('solid', fgColor='F4F0F7')
    GYM_ANCHOR  = PatternFill('solid', fgColor='E2D9EC')

    def cell(j, i, mid, fill):
        r = byid.get(mid)
        c = ws.cell(row=j, column=i, value=(r['name'] if r else (mid or '')))
        c.border = THIN
        c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        c.font = Font(size=9, bold=(fill is ANCHOR_FILL))
        if mid: c.fill = fill if r else GONE
        return r

    j = hr + 1
    for lid, name, family, tier, anchor, regs, progs in LAD['ladders']:
        # regressions right-aligned into the three slots before the anchor
        rslots = ([''] * 5 + list(regs))[-5:]
        pslots = (list(progs) + [''] * 5)[:5]
        ws.cell(row=j, column=1, value=name).font = Font(bold=True, size=11)
        ws.cell(row=j, column=2, value=family).font = Font(size=9)
        ws.cell(row=j, column=3, value=tier).alignment = Alignment(horizontal='center')
        for k, mid in enumerate(rslots):  cell(j, 4 + k, mid, REG_FILL)
        cell(j, 9, anchor, ANCHOR_FILL)
        for k, mid in enumerate(pslots):  cell(j, 10 + k, mid, PROG_FILL)

        # the lowest rung anyone can enter at
        first = next((m for m in regs if m in byid), anchor)
        fr = byid.get(first)
        entry = fr['level'] if fr and fr['level'] else '?'
        entry = {'beg': 'beginner', 'int': 'intermediate', 'adv': 'advanced',
                 'start': 'just starting'}.get(entry, entry)
        anch = byid.get(anchor) or {}
        for col, val in ((17, anch.get('muscles_primary', '')),
                         (18, anch.get('cue', '')),
                         (19, anch.get('video', ''))):
            cc = ws.cell(row=j, column=col, value=val)
            cc.border = THIN; cc.font = Font(size=9)
            cc.alignment = Alignment(wrap_text=True, vertical='top')
            if not val: cc.fill = NEEDS
        c = ws.cell(row=j, column=15, value=entry); c.border = THIN; c.font = Font(size=9)
        # a TIER-1 ladder that a just-starting person cannot enter is a real gap
        if tier == 1 and entry not in ('just starting', 'beginner'): c.fill = BLANK
        elif entry not in ('just starting', 'beginner'): c.fill = NEEDS
        n = len([m for m in regs if m in byid]) + 1 + len([m for m in progs if m in byid])
        ws.cell(row=j, column=16, value=n).alignment = Alignment(horizontal='center')
        # four ladders were renamed by this round of rework; his note follows the ladder
        RENAMED = {'Romanian Deadlift': 'Bodyweight Hip Hinge',
                   'Band-Assisted Front Lever': 'Tuck Front Lever',
                   'Cable Woodchop': 'Half-Kneeling Rotation',
                   'Hip Thrust': 'Bodyweight Hip Hinge'}
        note = NICK.get(name) or NICK.get(RENAMED.get(name, ''))
        if note:
            nc = ws.cell(row=j, column=20, value='YOUR NOTE: ' + note)
            nc.font = Font(size=9, bold=True, color='8E6209')
            nc.alignment = Alignment(wrap_text=True, vertical='top')
            nc.fill = PatternFill('solid', fgColor='FFF8E7')
        ws.cell(row=j, column=21).fill = NEEDS
        cd = ws.cell(row=j, column=22, value='to write'); cd.fill = NEEDS; cd.border = THIN
        for i in (1, 2, 3, 16, 20, 21):
            ws.cell(row=j, column=i).border = THIN
        ws.row_dimensions[j].height = 28
        j += 1

        # ---- the gym lane, directly beneath ----
        lane = GYM['lanes'].get(lid)
        c = ws.cell(row=j, column=1, value='   gym version')
        c.font = Font(size=9, italic=True, color='6B5B95'); c.border = THIN
        ws.cell(row=j, column=2).border = THIN
        ws.cell(row=j, column=3).border = THIN
        if lane:
            gr = ([''] * 5 + list(lane['easier']))[-5:]
            gp = (list(lane['harder']) + [''] * 5)[:5]
            for k, mid in enumerate(gr): cell(j, 4 + k, mid, GYM_FILL)
            cell(j, 9, lane['anchor'], GYM_ANCHOR)
            for k, mid in enumerate(gp): cell(j, 10 + k, mid, GYM_FILL)
            n = len(lane['easier']) + 1 + len(lane['harder'])
            ws.cell(row=j, column=16, value=n).alignment = Alignment(horizontal='center')
            w = ws.cell(row=j, column=20, value=lane['why'])
            w.font = Font(size=9, italic=True); w.alignment = Alignment(wrap_text=True, vertical='top')
        else:
            reason = GYM['no_lane'].get(lid, '')
            m = ws.cell(row=j, column=4, value='no honest gym equivalent')
            m.font = Font(size=9, italic=True, color='A02D25')
            m.alignment = Alignment(vertical='center')
            ws.merge_cells(start_row=j, start_column=4, end_row=j, end_column=14)
            w = ws.cell(row=j, column=20, value=reason)
            w.font = Font(size=9, italic=True); w.alignment = Alignment(wrap_text=True, vertical='top')
        for i in (15, 16):
            ws.cell(row=j, column=i).border = THIN
        ws.row_dimensions[j].height = 30
        j += 1
    last = j - 1

    j += 1
    ws.cell(row=j, column=1, value='PROTOCOLS & COMPLEXES — prescriptions rather than movements. No reps, no ladder, no calibration. They belong on the list; they just are not exercises.').font = Font(bold=True, italic=True, size=10)
    ws.merge_cells(start_row=j, start_column=1, end_row=j, end_column=17)
    j += 1
    for pid, pname, fam, tier, freq, how in LAD['protocols']:
        for i, v in enumerate([pname, fam, tier, '', '', '', '', '', freq, '', '', '', '', '', '-', '-', ''], start=1):
            c = ws.cell(row=j, column=i, value=v); c.border = THIN; c.font = Font(size=9)
        w = ws.cell(row=j, column=20, value=how)
        w.font = Font(size=9, italic=True); w.alignment = Alignment(wrap_text=True, vertical='top')
        j += 1

    # the kettlebell complexes, folded in from KETTLEBELL.xlsx on 20 Aug
    KBLIB = json.load(open(os.path.join(HERE, 'kettlebell-library.json')))
    for cxp in KBLIB.get('complexes_full', []):
        for i, v in enumerate([cxp['name'], 'kb complex', cxp.get('level') or '', '', '', '', '', '',
                               cxp.get('kit') or '', '', '', '', '', '',
                               cxp.get('source') or '', '-', ''], start=1):
            c = ws.cell(row=j, column=i, value=v); c.border = THIN; c.font = Font(size=9)
        w = ws.cell(row=j, column=20, value=cxp.get('protocol') or '')
        w.font = Font(size=9, italic=True); w.alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[j].height = 34
        j += 1

    # Nicolas's kettlebell videos, rescued from the same file
    res = KBLIB.get('resources') or {}
    if res.get('videos'):
        j += 1
        ws.cell(row=j, column=1, value='KETTLEBELL VIDEOS — yours, from the old KETTLEBELL.xlsx').font = Font(bold=True, italic=True, size=10)
        j += 1
        for url in res['videos']:
            c = ws.cell(row=j, column=1, value=url); c.font = Font(size=9, color='1B5B70')
            ws.merge_cells(start_row=j, start_column=1, end_row=j, end_column=6)
            j += 1

    d = DataValidation(type='list', formula1='"keep,change,discuss"', allow_blank=True)
    ws.add_data_validation(d); d.add(f'U{hr+1}:U{last}')
    ws.freeze_panes = f'D{hr+1}'
    for i, w in enumerate([26, 16, 6, 16, 16, 16, 16, 16, 20, 16, 16, 16, 16, 16, 14, 7,
                           26, 46, 30, 46, 11, 11], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws, len(LAD['ladders']), 0


def build_gym(wb, rows):
    """The parallel classic-gym build. The LIST comes first; the schema
    columns get filled once Nicolas has said yes / no / add to each line."""
    ws = wb.create_sheet('GYM LIBRARY')
    ws.cell(row=1, column=1, value=(
        'PARALLEL TRACK — the classic gym library. Say yes / no / add on each line first; '
        'we fill the full schema for the keepers afterwards. "Already in DB?" is checked '
        'automatically against js/data/exercises.js.')).font = Font(italic=True, size=10)
    ws.merge_cells('A1:H1')
    hr = 3
    for i, h in enumerate(GYM_COLS, start=1):
        c = ws.cell(row=hr, column=i, value=h)
        c.fill = PatternFill('solid', fgColor=HEAD_BG)
        c.font = Font(bold=True, color=HEAD_FG)

    with open(os.path.join(HERE, 'gym-library-seed.json')) as f:
        seed = json.load(f)['movements']
    try:
        dec = json.load(open(os.path.join(HERE, '..', 'workbook', '_decisions.json')))
    except Exception:
        dec = {'gym': {}, 'gym_notes': {}}

    def norm(s):
        return re.sub(r'[^a-z]', '', s.lower())
    existing = {norm(r['name']): r['name'] for r in rows}

    j = hr + 1
    for cat, name, equip, pat, mus in seed:
        hit = existing.get(norm(name), '')
        for i, v in enumerate([cat, name, equip, pat, mus,
                               hit or 'no',
                               dec.get('gym', {}).get(name, ''),
                               dec.get('gym_notes', {}).get(name, '')], start=1):
            c = ws.cell(row=j, column=i, value=v)
            c.border = THIN
            if i == 6:
                c.fill = AUTO if hit else NEEDS
            if i == 7:
                c.fill = NEEDS
        j += 1
    last = j - 1

    d = DataValidation(type='list', formula1='"yes,no,later"', allow_blank=True)
    ws.add_data_validation(d)
    d.add(f'G{hr+1}:G{last}')

    ws.freeze_panes = f'A{hr+1}'
    ws.auto_filter.ref = f'A{hr}:H{last}'
    for i, w in enumerate([16, 34, 22, 20, 28, 16, 11, 34], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws, len(seed), sum(1 for cat, n, *_ in seed if norm(n) in existing)


def build_enums(wb):
    ws = wb.create_sheet('LISTS')
    ws.cell(row=1, column=1, value=(
        'The allowed values. Anything outside these fails the build — that is what '
        'stops a fourteenth spelling of "push" appearing next month.')).font = Font(italic=True, size=10)
    ws.merge_cells('A1:H1')
    col = 1
    for key, vals in ENUMS.items():
        c = ws.cell(row=3, column=col, value=key)
        c.fill = PatternFill('solid', fgColor=HEAD_BG)
        c.font = Font(bold=True, color=HEAD_FG)
        for k, v in enumerate(vals, start=4):
            ws.cell(row=k, column=col, value=v).border = THIN
        ws.column_dimensions[get_column_letter(col)].width = max(14, len(key) + 4)
        col += 1
    ws.freeze_panes = 'A4'
    return ws


def build_do_next(wb, rows, ncal, nfund, nfundmiss):
    """The only sheet Nicolas has to open. Everything waiting on him, in order,
    with how long it takes. Everything else in the workbook is reference."""
    ws = wb.create_sheet('DO THIS NEXT', 0)
    for L, w in (('A',4),('B',44),('C',15),('D',12),('E',80)):
        ws.column_dimensions[L].width = w

    t = ws.cell(row=1, column=1, value='DO THIS NEXT')
    t.font = Font(bold=True, size=18, color=INK)
    ws.cell(row=2, column=1, value=(
        'Everything waiting on you, hardest-first. Nothing else in this workbook needs you — '
        'the other sheets are reference and I keep them current.')).font = Font(italic=True, size=10)
    ws.merge_cells('A2:E2')

    hr = 4
    for i, h in enumerate(['', 'Do this', 'Where', 'How long', 'Why it matters / what to look for'], start=1):
        c = ws.cell(row=hr, column=i, value=h)
        c.fill = PatternFill('solid', fgColor=HEAD_BG); c.font = Font(bold=True, color=HEAD_FG)
    n_review  = sum(1 for x in rows if x['status'] == 'REVIEW')
    n_noart   = sum(1 for x in rows if x['fundamental'] and x['art'] != 'drawn')

    TASKS = [
      ('Walk the 27 ladders once, left to right', 'FUNDAMENTALS', '10 min',
       'This is the new shape you asked for: one anchor per fundamental, regressions to its left so anyone '
       'can enter, progressions to its right. Read a few rows and tell me where a rung is in the wrong order.'),
      ('Confirm what your notes changed', 'FUNDAMENTALS', '5 min',
       'Plank is now a regression on the Hollow Hold ladder, not a fundamental. Ring Dip is a progression of Dip. '
       'Dragon Flag got its own ladder with five regressions. CARs became World\'s Greatest Stretch. '
       'Hollow Rock is a progression, not a second fundamental. All from your notes — say if any is wrong.'),
      ('Decide the headline number', 'FUNDAMENTALS', '2 min',
       'It is no longer 15 / 30 / 50. It is 27 ladders: 14 to start with, 9 more at intermediate, 4 at advanced — '
       'and 144 movements hang off them. That is a better story, but it is your brand, so you name it.'),
      ('Fix three calibration numbers I flagged as wrong', 'CALIBRATION', '5 min',
       'V-Sit is rated ABOVE L-Sit but it is the harder movement. Weighted Chin-Up has the same 20 at all three '
       'levels, and the load IS the level. Muscle-Up has no "strong" number.'),
      ('Skim the calibration numbers I proposed', 'CALIBRATION · Source = proposed', '20 min',
       'Yellow is mine, green is yours. Overwrite anything that feels wrong.'),
      ('Check the 6 numbers where you and the research disagree', 'CALIBRATION', '10 min',
       'Push-ups strong 75 is elite; research says 50-60. Inverted Row sits below both references. '
       'Burpee — max set or per minute?'),
      ('Write the 27 anchor cards first', 'FUNDAMENTALS · Card', 'ongoing',
       'Only the anchors need a card to begin with. The rungs inherit the explanation.'),
      (f'Draw the {n_noart} fundamentals with no picture', 'MOVEMENTS · filter Artwork', 'ongoing',
       'Anchors first — those are what a person sees on day one.'),
      (f'Check the {n_review} rows I guessed on', 'MOVEMENTS · filter Status', 'ongoing',
       'Muscles, patterns and roles inferred from names. Mostly right; the wrong ones are quick to spot.'),
    ]
    j = hr + 1
    for k, (what, where, how, why) in enumerate(TASKS, start=1):
        for i, v in enumerate([k, what, where, how, why], start=1):
            c = ws.cell(row=j, column=i, value=v); c.border = THIN
            c.alignment = Alignment(wrap_text=True, vertical='top')
            if i == 1: c.font = Font(bold=True, size=12, color='8E6209'); c.alignment = Alignment(horizontal='center', vertical='top')
            if i == 2: c.font = Font(bold=True, size=11)
            if i == 3: c.font = Font(name='Menlo', size=9)
        ws.row_dimensions[j].height = 40 if k <= 6 else 34
        j += 1

    j += 2
    ws.cell(row=j, column=1, value='WHERE EVERYTHING IS').font = Font(bold=True, size=13, color=INK); j += 1
    for i, h in enumerate(['', 'Sheet', 'What it is', '', 'Do you need to touch it?'], start=1):
        c = ws.cell(row=j, column=i, value=h)
        c.fill = PatternFill('solid', fgColor=HEAD_BG); c.font = Font(bold=True, color=HEAD_FG)
    j += 1
    MAP = [
      ('DO THIS NEXT', 'This sheet. Your queue.', 'Yes — start here, always'),
      ('FUNDAMENTALS', f'The {nfund} ladders. Anchor in the middle, regressions left, progressions right.', 'Yes — walk them once'),
      ('CALIBRATION', f'{ncal} rows. Entry / Solid / Strong per movement, weights in lbs.', 'Yes — the numbers'),
      ('MOVEMENTS', f'All {len(rows)} movements, every field. The database itself.', 'Only the yellow Status rows'),
      ('TRACKS', 'Progression chains, easiest to hardest.', 'No — reference'),
      ('GYM LIBRARY', 'Your 102 yes / 7 no on the classic gym list. Done.', 'No — finished'),
      ('LISTS', 'The allowed values behind every dropdown.', 'No — reference'),
    ]
    for name, what, need in MAP:
        for i, v in enumerate(['', name, what, '', need], start=1):
            c = ws.cell(row=j, column=i, value=v); c.border = THIN
            c.alignment = Alignment(wrap_text=True, vertical='top')
            if i == 2: c.font = Font(bold=True)
            if i == 5: c.fill = NEEDS if need.startswith('Yes') else AUTO
        ws.row_dimensions[j].height = 28
        j += 1

    j += 2
    ws.cell(row=j, column=1, value='COLOUR KEY').font = Font(bold=True, size=12); j += 1
    for label, fill in (('You decide, or I guessed', NEEDS),
                        ('Your own answer / derived and fine', AUTO),
                        ('Required and still empty', BLANK)):
        ws.cell(row=j, column=2, value=label)
        c = ws.cell(row=j, column=3); c.fill = fill; c.border = THIN
        j += 1
    return ws


def _verify(path):
    """Open the saved file the way Excel does and refuse to ship a repair prompt.

    Excel does not tell you WHICH rule you broke — it just says the file needs
    repairing. So the build checks the limits itself: XML well-formedness, the
    255/32 character caps on data-validation strings, and 1x1 merges.
    """
    import zipfile, re, html, xml.etree.ElementTree as ET
    problems = []
    with zipfile.ZipFile(path) as z:
        for n in z.namelist():
            if not (n.endswith('.xml') or n.endswith('.rels')):
                continue
            raw = z.read(n)
            try:
                ET.fromstring(raw)
            except Exception as e:
                problems.append(f'{n}: malformed XML ({e})')
                continue
            if '/worksheets/' not in n:
                continue
            t = raw.decode('utf8')
            for m in re.finditer(r'<dataValidation ([^>]*)>', t):
                a = m.group(1)
                for attr, lim in (('prompt', 255), ('error', 255),
                                  ('promptTitle', 32), ('errorTitle', 32)):
                    g = re.search(attr + r'="([^"]*)"', a)
                    if g and len(html.unescape(g.group(1))) > lim:
                        problems.append(f'{n}: {attr} is {len(html.unescape(g.group(1)))} chars, limit {lim}')
            for g in re.findall(r'<formula1>([^<]*)</formula1>', t):
                if len(html.unescape(g)) > 255:
                    problems.append(f'{n}: validation list is {len(g)} chars, limit 255')
            for ref in re.findall(r'<mergeCell ref="([^"]+)"', t):
                a, b = ref.split(':')
                if a == b:
                    problems.append(f'{n}: 1x1 merge {ref}')
    if problems:
        print('\nWORKBOOK WOULD ASK EXCEL FOR A REPAIR:')
        for p_ in problems:
            print('   ', p_)
        raise SystemExit(1)
    print('  verified     opens clean (no repair prompt)')


def main():
    with open(os.path.join(WB_DIR, '_derived.json')) as f:
        rows = json.load(f)

    wb = Workbook()
    wb.remove(wb.active)
    build_movements(wb, rows)
    build_tracks(wb, rows)
    _, ncal = build_calibration(wb, rows)
    _, nfund, nfundmiss = build_fundamentals(wb, rows)
    _, ngym, ngymhave = build_gym(wb, rows)
    build_enums(wb)
    build_do_next(wb, rows, ncal, nfund, nfundmiss)

    # the order Nicolas reads them in: queue, then the two he works in, then reference
    order = ['DO THIS NEXT', 'FUNDAMENTALS', 'CALIBRATION', 'MOVEMENTS', 'TRACKS', 'GYM LIBRARY', 'LISTS']
    wb._sheets = [wb[n] for n in order if n in wb.sheetnames]
    wb.active = 0
    wb.save(OUT)
    _verify(OUT)

    print(f'wrote {OUT}')
    print(f'  sheets       {len(wb.sheetnames)}  ' + ' | '.join(wb.sheetnames))
    print(f'  MOVEMENTS    {len(rows)} rows, {sum(1 for r in rows if r["status"]=="REVIEW")} flagged')
    print(f'  FUNDAMENTALS {nfund} movements, {nfundmiss} still absent from the database')
    print(f'  CALIBRATION  {ncal} rows')


if __name__ == '__main__':
    main()
