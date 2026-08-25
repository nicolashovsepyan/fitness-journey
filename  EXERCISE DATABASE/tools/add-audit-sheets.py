#!/usr/bin/env python3
"""
ADD THE TWO REVIEW SHEETS to EXERCISES.xlsx, in place.

  python3 "EXERCISE DATABASE/tools/add-audit-sheets.py"

Writes CALIBRATION AUDIT and FUNDAMENTALS PROPOSAL. Touches nothing else —
this edits the live workbook rather than regenerating it, so every answer
already in the file stays exactly where it is.
"""
import json, os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
WB   = os.path.join(HERE, '..', 'workbook', 'EXERCISES.xlsx')
REF  = json.load(open(os.path.join(HERE, 'calibration-reference.json')))['movements']
PROP = json.load(open(os.path.join(HERE, 'fundamentals-proposal.json')))['rows']
DEC  = json.load(open(os.path.join(HERE, '..', 'workbook', '_decisions.json')))['calibration']

HEAD = '16213E'
NEEDS = PatternFill('solid', fgColor='FFF2CC')
OK    = PatternFill('solid', fgColor='E8F5E9')
BAD   = PatternFill('solid', fgColor='FFE0E0')
WARN  = PatternFill('solid', fgColor='FFE8CC')
THIN  = Border(*[Side(style='thin', color='D8D8D8')]*4)

def num(v):
    if v in (None,'','n/a','N/A'): return None
    try: return float(str(v).strip())
    except ValueError: return None

def head(ws, row, cells, widths):
    for i,h in enumerate(cells, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = PatternFill('solid', fgColor=HEAD)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[row].height = 30
    for i,w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = load_workbook(WB)
for s in ('CALIBRATION AUDIT','FUNDAMENTALS PROPOSAL'):
    if s in wb.sheetnames: del wb[s]

# ---------------- CALIBRATION AUDIT ----------------
ws = wb.create_sheet('CALIBRATION AUDIT')
ws.cell(row=1, column=1, value=(
  'Your numbers against two independent references. RESEARCH = triangulated calisthenics sources '
  '(Calisthenic Movement, FitnessFAQs, Bodyweight Warrior, StrengthLevel, GMB, ATHLEAN-X). '
  'CSV = our own BENCHMARKS - Male.csv. Where they disagree with you, the reason is in the last column. '
  'Put yes/no in "Take it?" — nothing changes until you do.')).font = Font(italic=True, size=10)
ws.merge_cells('A1:K1')
head(ws, 3, ['Verdict','Movement','Unit','Yours','Research','Our CSV','I recommend','Take it?','Why'],
     [11, 28, 15, 13, 13, 13, 15, 10, 78])

def fmt(t): return '/'.join('-' if x is None else str(int(x)) for x in t)

rows=[]
for name, r in REF.items():
    mine = DEC.get(name)
    m = tuple(num(mine.get(k)) for k in ('beg','int','adv')) if mine else (None,None,None)
    res = tuple(r['research']) if r.get('research') else None
    csv = tuple(r['csv']) if r.get('csv') else None
    if not any(x is not None for x in m):
        rows.append(('blank', name, r['unit'], m, res, csv, None, r['note'])); continue
    worst=0.0
    for i in range(3):
        cand=[t[i] for t in (res,csv) if t]
        if not cand or m[i] is None: continue
        lo,hi=min(cand),max(cand)
        d = (lo-m[i])/max(lo,1) if m[i]<lo else ((m[i]-hi)/max(hi,1) if m[i]>hi else 0.0)
        worst=max(worst,d)
    v = 'ok' if worst<0.20 else ('check' if worst<0.45 else 'OFF')
    rec = None
    if v!='ok' and (res or csv):
        base = res or csv
        rec = tuple(int(round(sum(x[i] for x in (res,csv) if x)/len([x for x in (res,csv) if x]))) for i in range(3))
    rows.append((v, name, r['unit'], m, res, csv, rec, r['note']))

order={'OFF':0,'check':1,'ok':2,'blank':3}
rows.sort(key=lambda x:(order[x[0]], x[1]))

j=4
for v,name,unit,m,res,csv,rec,note in rows:
    vals=[v, name, unit, fmt(m), fmt(res) if res else '-', fmt(csv) if csv else '-',
          fmt(rec) if rec else ('-' if v!='blank' else 'fill it in'), '', note]
    for i,x in enumerate(vals, start=1):
        c=ws.cell(row=j,column=i,value=x); c.border=THIN
        if i==1:
            c.font=Font(bold=True, size=9)
            c.fill = BAD if v=='OFF' else WARN if v=='check' else OK if v=='ok' else NEEDS
        if i==8: c.fill=NEEDS
        if i==9: c.alignment=Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[j].height = 30
    j+=1
last=j-1
d=DataValidation(type='list', formula1='"yes,no"', allow_blank=True)
ws.add_data_validation(d); d.add(f'H4:H{last}')
ws.freeze_panes='C4'; ws.auto_filter.ref=f'A3:I{last}'

# ---------------- FUNDAMENTALS PROPOSAL ----------------
ws = wb.create_sheet('FUNDAMENTALS PROPOSAL')
ws.cell(row=1, column=1, value=(
  'My proposal against your current 50, so you can compare gut to research. Backed by the 7 primal movement '
  'patterns (squat, hinge, lunge, push, pull, ROTATION, gait). The 50 currently has no rotation and no '
  'frontal-plane movement at all. Sorted so everything I would change is at the top; "keep" rows are agreement.')
).font = Font(italic=True, size=10)
ws.merge_cells('A1:I1')
head(ws, 3, ['Verdict','#','Tier','Family','Currently','I propose','Alternatives','Your call','Why'],
     [13, 5, 6, 15, 34, 34, 34, 11, 74])

rank={'ADD':0,'SPLIT':1,'RECLASSIFY':2,'QUESTION':3,'keep':4}
j=4
for row in sorted(PROP, key=lambda r:(rank[r[0]], r[2], str(r[1]))):
    verdict, no, tier, family, cur, prop, alts, why = row
    for i,x in enumerate([verdict,no,tier,family,cur,prop,alts,'',why], start=1):
        c=ws.cell(row=j,column=i,value=x); c.border=THIN
        if i==1:
            c.font=Font(bold=True, size=9)
            c.fill = BAD if verdict=='ADD' else WARN if verdict in ('SPLIT','RECLASSIFY') else NEEDS if verdict=='QUESTION' else OK
        if i==8: c.fill=NEEDS
        if i in (5,6,7,9): c.alignment=Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[j].height = 32
    j+=1
last=j-1
d=DataValidation(type='list', formula1='"take yours,take mine,discuss"', allow_blank=True)
ws.add_data_validation(d); d.add(f'H4:H{last}')
ws.freeze_panes='B4'; ws.auto_filter.ref=f'A3:I{last}'

wb.save(WB)
print('added CALIBRATION AUDIT and FUNDAMENTALS PROPOSAL to EXERCISES.xlsx')
print('  audit rows      ', len(rows))
print('  proposal rows   ', len(PROP))
